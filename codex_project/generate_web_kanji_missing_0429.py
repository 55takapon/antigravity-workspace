import csv
from pathlib import Path

from openpyxl import load_workbook


SCRAPER_DIR = Path(r"C:\Users\hangy\.cursor\test\web-kanji-scraper")
OUTPUT = Path("web_kanji_missing_0429_rows.tsv")

PREFECTURES = [
    ("shizuoka", "静岡県"),
    ("miyagi", "宮城県"),
    ("ibaraki", "茨城県"),
    ("gifu", "岐阜県"),
    ("nagano", "長野県"),
    ("gunma", "群馬県"),
    ("tochigi", "栃木県"),
    ("fukushima", "福島県"),
    ("ishikawa", "石川県"),
    ("iwate", "岩手県"),
    ("aomori", "青森県"),
    ("yamanashi", "山梨県"),
    ("yamagata", "山形県"),
    ("toyama", "富山県"),
    ("akita", "秋田県"),
    ("fukui", "福井県"),
]

START_NUMBER = 7315
TIMESTAMP = "2026/04/29"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def build_rows():
    rows = []
    current_number = START_NUMBER
    per_prefecture = {}

    for slug, prefecture in PREFECTURES:
        workbook_path = SCRAPER_DIR / f"web_kanji_{slug}_companies.xlsx"
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [clean(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        header_index = {name: idx for idx, name in enumerate(headers)}

        count = 0
        for source_row in ws.iter_rows(min_row=2, values_only=True):
            if not any(clean(value) for value in source_row):
                continue

            def value(name):
                idx = header_index.get(name)
                return clean(source_row[idx]) if idx is not None and idx < len(source_row) else ""

            rows.append([
                str(current_number),
                prefecture,
                value("会社名"),
                value("代表"),
                value("URL"),
                value("お問い合わせページURL"),
                TIMESTAMP,
                value("送信NG"),
                value("検出キーワード"),
            ])
            current_number += 1
            count += 1

        per_prefecture[prefecture] = count

    return rows, per_prefecture


def main():
    rows, per_prefecture = build_rows()
    with OUTPUT.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t", lineterminator="\n")
        writer.writerows(rows)

    print(f"rows={len(rows)}")
    print(f"first={rows[0] if rows else ''}")
    print(f"last={rows[-1] if rows else ''}")
    print("per_prefecture=" + ", ".join(f"{k}:{v}" for k, v in per_prefecture.items()))


if __name__ == "__main__":
    main()
